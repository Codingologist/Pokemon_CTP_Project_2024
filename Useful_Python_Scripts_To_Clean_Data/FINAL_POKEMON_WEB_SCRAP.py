'''
Created on Sep 21, 2024

@author: FemiA

credits to Youtuber: 
ivangrov, which had provided me with a tutorial/base template 
to build this web scrapper
'''

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import shutil
import os
from selenium import webdriver
# selenium allows for autonomous on the web functions via. python
from selenium.webdriver.support import expected_conditions as EC
# provides a waiting time for selenium to operate
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException, NoSuchWindowException, NoAlertPresentException, UnexpectedAlertPresentException, TimeoutException, ElementClickInterceptedException, ElementNotInteractableException, WebDriverException
from webdriver_manager.firefox import GeckoDriverManager
# from webdriver_manager.chrome import ChromeDriverManager
import time 
from selenium.webdriver.support.select import Select
import random
import math
import pyautogui
import cv2
import tkinter as tk
import urllib.request
from selenium.webdriver.support.expected_conditions import _find_element
from selenium.webdriver.common.by import By
from typing import ItemsView
from pandas.core.frame import DataFrame
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import requests
import bs4
import time
import urllib
from bs4 import BeautifulSoup
 
row_Count = 2

File_a = 'Poke_STATS_O'
# intializes the excel sheet
workbook = load_workbook(filename="{Files_1}.xlsx".format(Files_1=File_a))
workbook.sheetnames

stop_count = 1
# Poke_Classes
sheet = workbook.active
pokemon_names = []

# iterates over a given Excel sheet's values and stores them in a list, so that it could iterated on/referenced later on

while  (sheet["A{r_values}".format(r_values=row_Count)].value != None):
    pokemon_names.append(sheet["A{r_values}".format(r_values=row_Count)].value)
    # print(pokemon_names)
    row_Count += 1
    
print("pokemon list", pokemon_names)

pokemon_name_count = 0
# pokemon_names = "Bulbasaur"
# creating a directory to save images
folder_name = "{pokemon_names}".format(pokemon_names=pokemon_names[pokemon_name_count])


def download_image(url, folder_name, num):

    # write image to file
# write image to file
    reponse = requests.get(url)
    if reponse.status_code == 200:
        with open(os.path.join(folder_name, str(num) + ".jpg"), 'wb') as file:
            file.write(reponse.content)


driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())

# for loop iterates over  list generated from the initial  Excel sheet.
for x in pokemon_names:
    folder_name = "{pokemon_names}".format(pokemon_names=pokemon_names[pokemon_name_count])
    # chromePath=r'C:\Users\net51\Documents\MyPythonScripts\Drivers\chromedriver.exe'
    # driver=webdriver.Chrome(chromePath)
    if not os.path.isdir(folder_name):
        os.makedirs(folder_name)

    search_URL = f"https://www.google.com/search?site=&tbm=isch&source=hp&biw=1873&bih=990&q={x}+ Pokemon"
    time.sleep(2)
    driver.get(search_URL)

    time.sleep(4)
    driver.execute_script("window.scrollTo(0, 0);")
    
    page_html = driver.page_source
    pageSoup = bs4.BeautifulSoup(page_html, 'html.parser')
    containers = pageSoup.findAll('div', {'class':"eA0Zlc WghbWd FnEtTd mkpRId m3LIae RLdvSe qyKxnc ivg-i PZPZlf GMCzAd"})
    # the div container main id goes here,. You will need this to obtain the # of containers in a given page
    
    print(len(containers))
    
    len_containers = len(containers)
    
    image_count = 60

    # for loop loops around the the google image pages until  the desired amount of images have been scrapped from image_count.
    for i in range(1, image_count + 1):
        
        # safeguard which pevents google from timing us out after searching around 25 images. 
        if i % 25 == 0:
            continue

        # once starting div_value path value has been obtained post path here, and make sure to modify the part where  the count changes, so you can interate through
        # will set up the program to be able to
    
        xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i)

        previewImageElement = driver.find_element_by_xpath(xPath)
        print(previewImageElement)
        previewImageURL = previewImageElement.get_attribute('src')
    
        # print(xPath)
    
        try:
            driver.find_element_by_xpath(xPath).click()
        except ElementNotInteractableException: 
            print("done")
            sys.exit()
    
        timeStarted = time.time()
      
        """these  long  lines of that essentially help the user not get timed out when and ideally extract the  xpath of a given target image so 
        that an image src link could be obtained, which would in turn help us to download a given Google image on the web"""
        try:
            imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
              
        except NoSuchElementException:
            time.sleep(2)
            print("evading getting stuck")
            xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 1)
            driver.find_element_by_xpath(xPath).click()
            try:
                imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
            except  NoSuchElementException:
                print("evading getting stuck")
                time.sleep(2)
                xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 3)
                driver.find_element_by_xpath(xPath).click()
                try:
                    imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
                except  NoSuchElementException:
                    time.sleep(4)
                    print("evading getting stuck")
                    time.sleep(2)
                    xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 10)
                    driver.find_element_by_xpath(xPath).click()
                    driver.find_element_by_xpath(xPath).click()
                    time.sleep(2)
                    imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
            
        except StaleElementReferenceException:
            time.sleep(2)
            print("evading getting stuck")
            xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 1)
            driver.find_element_by_xpath(xPath).click()
            imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
     
        # attemps to obtain the src link, of a given image so that it could be readily downlaoded
        try:
            imageURL = imageElement.get_attribute('src')
        except StaleElementReferenceException:
            time.sleep(2)
            print("evading getting stuck")
            xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 1)
            driver.find_element_by_xpath(xPath).click()
            
        print(imageURL)

        # attempts to download a given image into this directory, if it fails, it will move onto the next given image to scrap by starting the 
        # for loop again
        try:
            download_image(imageURL, folder_name, i)
            print("Downloaded element %s out of %s total. URL: %s" % (i, len_containers + 1, imageURL))
            
        except:
            print("Couldn't download an image %s, continuing downloading the next one" % (i))
            
            """time.sleep(2)
            print("evading getting stuck")
            xPath = '/html/body/div[3]/div/div[15]/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div[{div_value}]'.format(div_value=i + 1)
            driver.find_element_by_xpath(xPath).click()
            imageElement = driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div/div/div/c-wiz/div/div[2]/div[2]/div/div[2]/c-wiz/div/div[3]/div[1]/a/img[1]')
            imageURL = imageElement.get_attribute('src')
            print(imageURL)
            download_image(imageURL, folder_name, i)
            print("Downloaded element %s out of %s total. URL: %s" % (i, len_containers + 1, imageURL))"""
        
    pokemon_name_count += 1
