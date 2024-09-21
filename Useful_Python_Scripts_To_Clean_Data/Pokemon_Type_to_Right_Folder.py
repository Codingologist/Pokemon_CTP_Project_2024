'''
Created on Sep 21, 2024

@author: FemiA
'''

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import shutil
import os

File_a='Poke_STATS_O'


 
row_Count= 2

workbook = load_workbook(filename="{Files_1}.xlsx".format(Files_1=File_a))


workbook.sheetnames


#Poke_Classes

sheet = workbook.active

list_pokemon= []
list_of_pokemon_types =[]
dict_poke = {}
poke_count =0

while  (sheet["A{r_values}".format(r_values=row_Count)].value != None):
    list_pokemon.append(sheet["A{r_values}".format(r_values=row_Count)].value)
    
    list_of_pokemon_types.append(sheet["I{r_values}".format(r_values=row_Count)].value)

    
    row_Count +=1
    
print(list_pokemon)
print(list_of_pokemon_types)


for x in list_pokemon:
    dict_poke[x] =list_of_pokemon_types[poke_count]
    poke_count += 1
    

print(dict_poke)
poke_count =0
row_Count =2

# Get the list of all files and directories
path = "Pokemon_Images"
dir_list = os.listdir(path)
print("Files and directories in '", path, "' :")
# prints all files
#print(dir_list)
#print(len(dir_list))



# Get the list of all files and directories
path_2= "Poke_Classes"
dir_list_2 = os.listdir(path_2)
print("Files and directories in '", path_2, "' :")
# prints all files
print(dir_list_2)
#print(len(dir_list))    

directory = path_2
parent_dir =path
path = os.path.join(parent_dir, directory)  
pokemon_missing= []

for x in dir_list:
    #print(x)
    print(poke_count)
    #print(list(dict_poke.keys())[poke_count])
    #print(dict.get('pokename').format(pokename=x))
    for y in dict_poke:
        try:
            if x == list(dict_poke.keys())[poke_count]:
                print("Working")
                print(list(dict_poke.keys())[poke_count])
                # path to source directory
                src_dir = 'Pokemon_Images/{name}'.format(name=x)
                dest_dir = 'Poke_Classes/{name_2}/{name}'.format(name_2= dict_poke.get(x),name=x )
                #files = os.listdir(src_dir)
                shutil.copytree(src_dir, dest_dir)  
                poke_count=0
                row_Count +=1
                break
        except IndexError:
            print("pokemon not here")
            #print(x)
            sheet["J{r_values}".format(r_values=row_Count)] = x
            pokemon_missing.append(x)
            row_Count +=1
            poke_count=0
           
        else:
            poke_count +=1
            sheet["J{r_values}".format(r_values=row_Count)]= x
            row_Count +=1
            pokemon_missing.append(x)
            
#print(pokemon_missing)
workbook.save(filename="{Files_1}.xlsx".format(Files_1 = File_a))               

 



