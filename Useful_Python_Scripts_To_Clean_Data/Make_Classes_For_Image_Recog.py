'''
Created on Sep 21, 2024

@author: FemiA
'''
'''
Created on Sep 19, 2024

@author: FemiA
'''

from openpyxl import Workbook
from openpyxl import load_workbook
import sys
File_a='Poke_STATS_2'


 
row_Count= 2

workbook = load_workbook(filename="{Files_1}.xlsx".format(Files_1=File_a))


workbook.sheetnames


#Poke_Classes

sheet = workbook.active

list_of_pokemon_types= []

while  (sheet["A{r_values}".format(r_values=row_Count)].value != None):
    list_of_pokemon_types.append(sheet["I{r_values}".format(r_values=row_Count)].value)
    print(list_of_pokemon_types)
    row_Count +=1
    
 
 
 
import os
 
root_path = 'Poke_Classes'
 
for items in list_of_pokemon_types:
    path = os.path.join(root_path, items)
    os.mkdir(path)   
    
    
                                 
                                 
    
    
    
